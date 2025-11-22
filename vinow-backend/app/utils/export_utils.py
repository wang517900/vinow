商家板块6财务中心
import os
import uuid
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any
import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.models.finance import ReportType


class ExportUtils:
    """导出工具类"""
    
    @staticmethod
    def ensure_export_dir() -> str:
        """确保导出目录存在"""
        try:
            export_dir = os.path.join(os.getcwd(), settings.upload_dir)
            os.makedirs(export_dir, exist_ok=True)
            return export_dir
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"创建导出目录失败: {str(e)}"
            )
    
    @staticmethod
    def generate_filename(merchant_id: str, report_type: ReportType, format: str) -> str:
        """生成文件名"""
        # 验证并清理输入参数
        if not merchant_id or not format:
            raise ValueError("商户ID和文件格式不能为空")
        
        # 清理商户ID中的非法字符
        clean_merchant_id = re.sub(r'[^\w\-_]', '_', str(merchant_id))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{clean_merchant_id}_{report_type.value}_report_{timestamp}.{format.lower()}"
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "财务数据"
    ) -> str:
        """导出数据到Excel"""
        try:
            export_dir = ExportUtils.ensure_export_dir()
            filepath = os.path.join(export_dir, filename)
            
            # 处理空数据情况
            if not data:
                # 创建空的Excel文件
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                ws.cell(row=1, column=1, value="无数据")
                wb.save(filepath)
                return filepath
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 使用openpyxl创建带格式的Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 添加表头
            headers = list(data[0].keys()) if data else []
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=str(header))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 添加数据
            for row, item in enumerate(data, 2):
                for col, key in enumerate(headers, 1):
                    value = item.get(key, '')
                    # 处理特殊数据类型
                    if isinstance(value, (dict, list)):
                        value = str(value)
                    ws.cell(row=row, column=col, value=value)
            
            # 自动调整列宽
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    if max_length > 0:
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                # 列宽调整失败不影响主要功能
                pass
            
            wb.save(filepath)
            return filepath
            
        except PermissionError:
            raise HTTPException(
                status_code=500,
                detail=f"没有权限写入文件: {filepath}"
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Excel导出失败: {str(e)}"
            )
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str) -> str:
        """导出数据到CSV"""
        try:
            export_dir = ExportUtils.ensure_export_dir()
            filepath = os.path.join(export_dir, filename)
            
            # 处理空数据情况
            if not data:
                # 创建空的CSV文件
                with open(filepath, 'w', encoding='utf-8-sig') as f:
                    f.write("")
                return filepath
            
            df = pd.DataFrame(data)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            return filepath
            
        except PermissionError:
            raise HTTPException(
                status_code=500,
                detail=f"没有权限写入文件: {filepath}"
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"CSV导出失败: {str(e)}"
            )
    
    @staticmethod
    def get_file_url(filepath: str) -> str:
        """获取文件访问URL"""
        try:
            filename = os.path.basename(filepath)
            # URL安全检查
            if '..' in filename or '/' in filename or '\\' in filename:
                raise ValueError("无效的文件名")
            return f"/api/merchant/finances/reports/download/{filename}"
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"生成文件URL失败: {str(e)}"
            )
    
    @staticmethod
    def calculate_file_size(filepath: str) -> int:
        """计算文件大小"""
        try:
            if os.path.exists(filepath):
                return os.path.getsize(filepath)
            return 0
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"计算文件大小失败: {str(e)}"
            )
    
    @staticmethod
    def get_expires_time(hours: int = 24) -> datetime:
        """获取过期时间"""
        if hours <= 0:
            raise ValueError("过期时间必须大于0小时")
        return datetime.now() + timedelta(hours=hours)
    
    @staticmethod
    def delete_file(filepath: str) -> bool:
        """删除文件"""
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
                return True
            return False
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"删除文件失败: {str(e)}"
            )
    
    @staticmethod
    def validate_format(format: str) -> bool:
        """验证文件格式是否支持"""
        supported_formats = ['xlsx', 'csv']
        return format.lower() in supported_formats